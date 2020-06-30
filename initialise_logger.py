from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, colors, PatternFill, Font, Fill
import shelve
from datetime import date
from file_paths import *  # paths for the different files to use


def stylise_cells(ws, cell_range, allign_style='center'):
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    ft = Font(color='000000', bold=True, name='Times New Roman')
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal=allign_style, vertical=allign_style)
            cell.font = ft


def generate_times(start='7:00', end='24:00'):
    """
    Generates times from any start to end with a 15 minute interval.
    """
    start_hour, start_min = start.split(':')
    start_hour = int(start_hour)
    times = []
    times.append('{:0>2d}:{}'.format(start_hour, start_min))
    minutes = {'00': '15', '15': '30', '30': '45', '45': '00'}
    while times[-1] != end:
        while times[-1].split(':')[1] != '45':
            times.append('{:0>2d}:{}'.format(start_hour, minutes[start_min]))
            start_min = minutes[start_min]
        start_hour += 1
        start_min = minutes[start_min]
        times.append('{:0>2d}:00'.format(start_hour))
    return times


def initialise_sheet(times):
    """
    Add times to a column in the sheet
    """
    # Enter your destination file - i.e where the excel file is to be saved.
    dest_file = log_file
    wb = Workbook()
    sheet = wb.active
    stylise_cells(sheet, 'B2:B70')
    sheet['B2'] = 'TIME'
    curr_column, curr_row = 'B', 3
    for time_data in times[:-1]:
        cell_name = f'{curr_column}{curr_row}'
        sheet[cell_name] = time_data
        curr_row += 1
    wb.save(dest_file)


def shelve_data(times):
    """
    Initialise the various arrays needed for the program and shelves them for
    use in the main logger.
    """
    data = {}
    activities = [None, 'Golf', 'Gardening', 'Studying', 'Movie', 'Phone']
    act_data = [0 for _ in range(len(times))]
    added = [False for _ in range(len(times))]
    merged = []
    date_today = date.today()
    colours_list = ['FFFFFF', '3E732F', 'C1C483', 'D93A28', '21384E', '26558B',
                    '6C693C', 'F55995', 'A42D41', 'D18D0E', 'F16102', '6C823F', 'E3A002']
    data['activity_names'] = activities
    data['act_data'] = act_data
    data['date'] = date_today
    data['added_sheet'] = added
    data['merged_cells'] = merged
    data['colours'] = colours_list
    data['time_list'] = times
    shelve_file = shelve_path
    with shelve.open(shelve_file) as fhand:
        fhand['data'] = data


def main():
    time_list = generate_times()
    initialise_sheet(time_list)
    shelve_data(time_list)


if __name__ == '__main__':
    main()
