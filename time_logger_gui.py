from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, colors, PatternFill, Font, Fill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import os
import sys
import shelve
from file_paths import *  # paths for the different files to use
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.backend_bases import key_press_handler
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from PIL import ImageGrab


class TimeLogger:
    def __init__(self, master):
        """
        Initialise all the files - path, date.
        Set up the Tk windows and buttons etc
        """
        # path must be same as the initialise logger
        self.file_path = log_file  # in file_paths program
        self.shelve_file = shelve_path
        self.date_today = date.today()
        self.d1 = self.date_today.strftime("%d-%m-%Y")
        self.time_list, self.activities, self.act_data, self.added_sheet, self.merged_cells, self.colours = [], [], [], [], [], []
        self.start_row = 2
        self.border = Border(left=Side(border_style='thin', color='000000'),
                             right=Side(border_style='thin', color='000000'),
                             top=Side(border_style='thin', color='000000'),
                             bottom=Side(border_style='thin', color='000000'))
        self.ft = Font(color='FFFFFF', bold=True, name='Times New Roman')
        self.allign_style = 'center'
        self.get_info()
        wb = load_workbook(self.file_path)
        ws = wb.active
        self.st_column = self.get_column(ws)  # figure out which column is to be appended today
        date_cell = '{}{}'.format(self.st_column, self.start_row)
        ws[date_cell] = self.d1
        self.cell_range = '{}{}:{}{}'.format(
            self.st_column, self.start_row, self.st_column, self.start_row + len(self.time_list)-1)
        self.stylise_cells(ws)
        wb.save(self.file_path)

        # tkinter setup
        self.master = master
        self.master.protocol("WM_DELETE_WINDOW", self.store_info)
        self.win_width = 650
        self.win_height = 630
        self.master.geometry("650x630")
        self.tabs = ttk.Notebook(self.master)
        self.tab1 = ttk.Frame(self.tabs)
        self.tab2 = ttk.Frame(self.tabs)
        self.tab3 = ttk.Frame(self.tabs)
        self.tabs.add(self.tab1, text="Time Logger")
        self.tabs.add(self.tab2, text="Manage Activities")
        self.tabs.add(self.tab3, text="Data Analytics")
        self.tabs.pack(expand=1, fill='both')
        self.blue = "#26558B"
        self.red = "#A42D41"
        self.white = "#F8F8FF"

        # tab1
        self.logger_frame = Frame(self.tab1, width=650, height=650, bg=self.blue)
        self.logger_frame.pack()

        self.title = Label(self.tab1, width=25, text="TIME LOGGER", font=(
            'System', 30, 'bold'), bg=self.blue, fg=self.white)
        self.title.place(relx=0, rely=0.04)

        self.button_texts = []
        self.buttons = []
        self.gen_radiobuttons()

        self.instructions = StringVar(
            value='Select activity, then add times in hh:mm format.')
        self.instruct_label = Label(self.tab1, textvariable=self.instructions,
                                    bg=self.blue, fg=self.white, anchor=CENTER, font=("System", 18, 'bold'))
        self.instruct_label.place(relx=0.02, rely=0.68)

        self.start_label = Label(self.tab1, text='START', bg=self.blue,
                                 fg=self.white, font=("System", 18, 'bold'))
        self.end_label = Label(self.tab1, text='END', bg=self.blue,
                               fg=self.white, font=("System", 18, 'bold'))
        self.start_label.place(relx=0.02, rely=0.77)
        self.end_label.place(relx=0.36, rely=0.77)

        self.start_text = StringVar(value='07:00')
        self.end_text = StringVar(value='24:00')
        self.start_entry = Entry(self.tab1, textvariable=self.start_text, fg=self.blue, bg=self.white, font=(
            'System', 18, 'bold'), justify=LEFT, width=6)
        self.end_entry = Entry(self.tab1, textvariable=self.end_text, fg=self.blue, bg=self.white, font=(
            'System', 18, 'bold'), justify=LEFT, width=6)
        self.start_entry.place(relx=0.19, rely=0.77)
        self.end_entry.place(relx=0.49, rely=0.77)
        self.submit_time = Button(self.tab1, text="UPDATE", bg=self.red,
                                  fg=self.white, command=self.input_time, font=(
                                      'System', 18, 'bold'))
        self.submit_time.place(relx=0.72, rely=0.76)
        self.master.bind('<Return>', self.input_time)

        self.analyse = Button(self.tab1, text="ANALYSE", width=11, bg=self.red, fg=self.white, command=lambda: self.change_tab(2),  font=(
            'System', 18, 'bold'))
        self.edit_activities = Button(self.tab1, text="EDIT DATA", width=11, bg=self.red, fg=self.white, command=lambda: self.change_tab(1),  font=(
            'System', 18, 'bold'))
        self.quit_program = Button(self.tab1, text="QUIT", width=11, bg=self.red, fg=self.white, command=self.store_info, font=(
            'System', 18, 'bold'))
        self.analyse.place(relx=0.03, rely=0.88)
        self.edit_activities.place(relx=0.35, rely=0.88)
        self.quit_program.place(relx=0.67, rely=0.88)

        # tab2
        self.edit_frame = Frame(self.tab2, width=650, height=650, bg=self.blue)
        self.edit_frame.pack()
        self.edit_title = Label(self.tab2, width=25, text="ACTIVITIES", font=(
            'System', 30, 'bold'), bg=self.blue, fg=self.white)
        self.edit_title.place(relx=0, rely=0.03)

        self.data_text = Text(self.tab2, font=('System', 18, 'bold'),
                              height=9, width=32, bg=self.white, fg=self.blue)
        self.data_text.place(relx=0.075, rely=0.14)
        self.data_text.configure(state='disabled')

        self.add_act = StringVar(value='Enter activity to add.')
        self.add_entry = Entry(self.tab2, textvariable=self.add_act, font=(
            'System', 18, 'bold'), bg=self.white, fg=self.blue, width=25)
        self.del_act = StringVar(value='Enter activity to remove.')
        self.del_entry = Entry(self.tab2, textvariable=self.del_act, font=(
            'System', 18, 'bold'), bg=self.white, fg=self.blue, width=25)
        self.add_entry.place(relx=0.075, rely=0.67)
        self.del_entry.place(relx=0.075, rely=0.79)

        self.submit_add = Button(self.tab2, text='SUBMIT', bg=self.red, fg=self.white, font=(
            'System', 18, 'bold'), command=lambda: self.edit_activity(0))
        self.submit_del = Button(self.tab2, text='SUBMIT', bg=self.red, fg=self.white, font=(
            'System', 18, 'bold'), command=lambda: self.edit_activity(1))
        self.submit_add.place(relx=0.75, rely=0.66)
        self.submit_del.place(relx=0.75, rely=0.78)

        self.analyse_change = Button(self.tab2, text="ANALYSE", width=11, bg=self.red, fg=self.white, command=lambda: self.change_tab(2),  font=(
            'System', 18, 'bold'))
        self.go_log = Button(self.tab2, text="EDIT LOGS", width=11, bg=self.red, fg=self.white, command=lambda: self.change_tab(0),  font=(
            'System', 18, 'bold'))
        self.quit_fn = Button(self.tab2, text="QUIT", width=11, bg=self.red, fg=self.white, command=self.store_info, font=(
            'System', 18, 'bold'))
        self.analyse_change.place(relx=0.03, rely=0.89)
        self.go_log.place(relx=0.35, rely=0.89)
        self.quit_fn.place(relx=0.67, rely=0.89)
        self.show_activity()

        # tab3
        self.data_frame = Frame(self.tab3, width=650, height=630, bg=self.blue)
        self.data_frame.pack()
        self.data_title = Label(self.tab3, width=25, text="DATA ANALYSIS", font=(
            'System', 30, 'bold'), bg=self.blue, fg=self.white)
        self.data_title.place(relx=0, rely=0.03)

        self.canvas_chart = Canvas(self.tab3, width=450, height=300)
        self.canvas_chart.place(relx=0.14, rely=0.13)

        self.data_analysed = Text(self.tab3, width=35, height=4, font=(
            'System', 18, 'bold'), bg=self.blue, fg=self.white)
        self.data_analysed.place(relx=0.06, rely=0.66)
        self.data_analysed.insert(INSERT, 'Hit analyse to see the magic ;)')
        self.data_analysed.configure(state='disabled')

        self.show_analysis = Button(self.tab3, text='ANALYSE', width=11, command=self.anaylyse_data,
                                    fg=self.white, bg=self.red, font=('System', 18, 'bold'))
        self.save_analysis = Button(self.tab3, text='SAVE',  width=11, command=self.save_data,
                                    fg=self.white, bg=self.red, font=('System', 18, 'bold'))
        self.end_prg = Button(self.tab3, text='QUIT',  width=11, command=self.store_info,
                              fg=self.white, bg=self.red, font=('System', 18, 'bold'))
        self.show_analysis.place(relx=0.03, rely=0.89)
        self.save_analysis.place(relx=0.35, rely=0.89)
        self.end_prg.place(relx=0.67, rely=0.89)

    def change_tab(self, n):
        """
        Changing the tab
        """
        self.tabs.select(n)

    def gen_radiobuttons(self):
        """
        Creates the radiobuttons for the multiple choice in a for loop with max
        12 activities.
        """

        self.option = IntVar(value=1)
        acts = self.activities[1:]
        max = len(acts)
        x, y = 0.1, 0.18
        for i in range(12):
            if i < max:
                curr_act = acts[i].upper()
            else:
                curr_act = '(Add activity)'
            if i == 6:
                x = 0.64
                y = 0.18
            act_text = StringVar()
            act_text.set(curr_act)
            self.button_texts.append(act_text)
            val = i+1
            btn_choice = Radiobutton(self.tab1, textvariable=act_text,
                                     variable=self.option, value=val, bg=self.blue, fg=self.white, selectcolor=self.red, font=("System", 17, "bold"))
            btn_choice.place(relx=x, rely=y)
            self.buttons.append(btn_choice)
            if i >= max:
                btn_choice['state'] = DISABLED
            y += 0.08

    def get_info(self):
        """
        Retrieves all the information from the shelf, checks the shelf date.
        """
        data = {}
        with shelve.open(self.shelve_file) as fhand:
            data = fhand['data']
        self.time_list = data['time_list']
        self.colours = data['colours']
        self.activities = data['activity_names']
        if self.date_today == data['date']:  # same day restart
            self.act_data = data['act_data']
            self.added_sheet = data['added_sheet']
            self.merged_cells = data['merged_cells']
        else:
            self.act_data = [0 for _ in range(len(self.time_list))]
            self.added_sheet = [False for _ in range(len(self.time_list))]

    def store_info(self):
        """
        Stores all the info into the shelf - called by the quit or X button - and
        quits the program -  Edit the protocol.
        """
        data = {}
        data['activity_names'] = self.activities
        data['act_data'] = self.act_data
        data['date'] = self.date_today
        data['added_sheet'] = self.added_sheet
        data['merged_cells'] = self.merged_cells
        data['colours'] = self.colours
        data['time_list'] = self.time_list
        with shelve.open(self.shelve_file) as fhand:
            fhand['data'] = data
        sys.exit()

    def input_time(self, event=None):
        """
        Inputs the working time from the tk modules - checks for conflict with
        other data already stored, if conflict calls upon the conflict manager,
        and if yes then calls upon update info. Also validates the input.
        """
        start_time = self.start_text.get()
        end_time = self.end_text.get()
        act_choice = self.option.get()
        valid, conflict = self.validate_input(start_time, end_time, act_choice)
        if not valid:
            self.instructions.set('Invalid input. Please re-enter.')
        else:
            wb = load_workbook(self.file_path)
            ws = wb.active
            if conflict:
                ok = messagebox.askokcancel(
                    title='Conflict', message='The data you entered conflicts with a log already present. Overwrite logs?')
                if ok:
                    self.conflict_manager(ws)
                else:
                    wb.save(self.file_path)
                    return
            self.update_info(start_time, end_time, act_choice)
            self.update_log(ws)
            wb.save(self.file_path)
            self.instructions.set("Log saved. Enter next log.")
            self.master.after(1500, self.reinitialies_values)

    def validate_input(self, s, e, ch):
        """
        Validates the input times and returns a boolean.
        """
        if len(s) != 5 or len(e) != 5:
            return False, None
        if s == e:
            return False, None
        if s not in self.time_list or e not in self.time_list:
            return False, None
        # at this point, times are valid - have to check for conflicts
        start_index = self.time_list.index(s)
        end_index = self.time_list.index(e)
        if not start_index < end_index:
            return False, None
        for i in range(start_index, end_index):
            if self.act_data[i] not in [0, ch]:
                return True, True
        return True, False  # no conflict

    def update_info(self, start, end, opt):
        """
        Called by the input time function, edits the act_data as per required.
        """
        start_index, end_index = self.time_list.index(start), self.time_list.index(end)
        for i in range(start_index, end_index):
            self.act_data[i] = opt

    def conflict_manager(self, ws):
        """
        Unmerge all the cells and change the added list to
        include everything.
        """
        merged_data = self.merged_cells[::]
        for merged in merged_data:
            ws.unmerge_cells(merged)
            self.merged_cells.remove(merged)

        # so everything get's rechecked
        self.clear_column(ws)

    def update_log(self, ws):
        """
        Add the information to the spreadsheet and do the merging etc
        """
        curr_row = self.start_row + 1
        for i in range(len(self.time_list)):
            if not self.added_sheet[i]:
                curr_cell = '{}{}'.format(self.st_column, curr_row)
                act_id = self.act_data[i]
                activity = self.activities[act_id]
                if activity is not None:
                    if i < len(self.act_data) and act_id == self.act_data[i+1]:
                        self.added_sheet[i+1] = True
                        t = i + 1
                        while t < len(self.act_data)-1 and self.act_data[t] == self.act_data[t+1]:
                            self.added_sheet[t+1] = True
                            t += 1
                        end_row = curr_row + t - i
                        end_cell = '{}{}'.format(self.st_column, end_row)
                        to_merge = '{}:{}'.format(curr_cell, end_cell)
                        self.merged_cells.append(to_merge)
                        ws.merge_cells(to_merge)
                    cell = ws[curr_cell]
                    cell.value = activity.upper()
                    cell.font = Font(color=colors.WHITE)
                    c = self.colours[act_id]
                    cell.fill = PatternFill(fill_type='solid', start_color=c, end_color=c)
                    self.added_sheet[i] = True
            curr_row += 1

    def clear_column(self, ws):
        """
        Clears the column in case of any conflict.
        """
        self.added_sheet = [False for _ in range(len(self.time_list))]
        curr_row = self.start_row + 1
        curr_cell = '{}{}'.format(self.st_column, curr_row)
        for i in range(len(self.time_list)):
            cell = ws[curr_cell]
            cell.value = ''
            cell.border = self.border
            cell.fill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')
            cell.alignment = Alignment(horizontal=self.allign_style, vertical=self.allign_style)
            curr_row += 1
            curr_cell = '{}{}'.format(self.st_column, curr_row)

    def get_column(self, ws):
        """
        Runs from init, checks the excel sheet for which
        column should be used for today.
        """
        rows = ws.iter_rows(min_row=self.start_row,
                            max_row=self.start_row)  # the row where your headings are
        row = next(rows)
        headings = [c.value for c in row]
        col_letter = ''
        for col, heading in enumerate(headings):
            if not col:
                continue  # skip the first col
            if heading == self.d1 or heading is None:
                col_letter = get_column_letter(col+1)
                break
        if not col_letter:
            col_letter = get_column_letter(len(headings)+1)
        return col_letter

    def stylise_cells(self, ws):
        """
        Called upon by init after getting start column to stylise the rows
        """
        rows = ws[self.cell_range]
        for row in rows:
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(horizontal=self.allign_style, vertical=self.allign_style)

    def edit_activity(self, n):
        """
        Adds an activity to the activity list and then appends it to the
        checkbox/radiobutton.
        """
        if not n:  # adding an activity
            act_to_add = self.add_act.get().capitalize()
            if act_to_add not in self.activities:
                if not len(self.activities) < 13:
                    self.add_act.set('Too many activities. Delete then add.')
                else:
                    self.activities.append(act_to_add)
                    self.add_act.set('Activity added! Enter another.')
            else:
                self.add_act.set('Activity already exists.')
        else:
            act_to_del = self.del_act.get().capitalize()
            if act_to_del in self.activities:
                if messagebox.askokcancel(title='Warning', message='Deleting an activity will remove all logged info for that activity.'):
                    ind = self.activities.index(act_to_del)
                    self.activities.pop(ind)
                    for i in range(len(self.act_data)):
                        if self.act_data[i] < ind:
                            continue
                        elif self.act_data[i] == ind:
                            self.act_data[i] = 0
                        else:
                            self.act_data[i] -= 1
                    wb = load_workbook(self.file_path)
                    ws = wb.active
                    self.conflict_manager(ws)
                    self.update_log(ws)
                    wb.save(self.file_path)
                    self.del_act.set("Activity deleted. Enter another.")

            else:
                self.del_act.set('Activity does not exist. Check spelling.')
        self.show_activity()
        self.edit_radiobuttons()
        self.master.after(1500, self.reinitialies_values)

    def show_activity(self):
        """
        Displays the activities that are stored to the activity text.
        """
        self.data_text.configure(state='normal')
        self.data_text.delete('1.0', END)
        str_to_show = ''
        if not len(self.activities) > 1:
            str_to_show = 'Add an activity to see it here.'
        else:
            for i, act in enumerate(self.activities):
                if not i:
                    continue
                str_to_show += "{:0>2d}. {}\n".format(i, act)
        self.data_text.delete('1.0', END)
        self.data_text.insert(INSERT, str_to_show)
        self.data_text.configure(state='disabled')

    def edit_radiobuttons(self):
        """
        Edit the radiobutton when activities are added.
        """
        acts = self.activities[1:]
        max = len(acts)
        for i in range(len(self.button_texts)):
            if i < max:
                self.button_texts[i].set(acts[i])
                self.buttons[i]['state'] = NORMAL
            else:
                self.button_texts[i].set('(Add activity)')
                self.buttons[i]['state'] = DISABLED

    def reinitialies_values(self):
        self.start_text.set("hh:mm")
        self.end_text.set("hh:mm")
        self.instructions.set("Enter times below and add to log.")
        self.add_act.set("Add activity.")
        self.del_act.set("Delete activity.")

    def anaylyse_data(self):
        """
        Call upon matplotlib. generate the pie - also add the time logged for the
        day and the most logged activity.
        """
        start, end = self.time_list[0].split(':'), self.time_list[-1].split(':')
        s = int(start[0]) if start[1] == '00' else int(start[0]) + int(start[1])/60
        e = int(end[0]) if end[1] == '00' else int(end[0]) + int(end[1])/60
        hours = int(e-s)
        min = int((e - s - hours)*60)
        total_hours = '{:0>2d}:{:0>2d}'.format(hours, min)
        slots_tracked = len(self.act_data) - self.act_data.count(0)
        h, min = divmod(slots_tracked*15, 60)
        hours_tracked = '{:0>2d}:{:0>2d}'.format(h, min)
        labels_graph = self.activities[1:]
        values, labels = [], []
        for i in range(len(labels_graph)):
            if self.act_data.count(i+1) > 0:
                values.append(self.act_data.count(i+1))
                labels.append(labels_graph[i])
        # print(labels)
        # print(values)
        max_id = values.index(max(values))
        max_act = labels[max_id]
        max_h, max_m = divmod(values[max_id]*15, 60)
        max_tracked = '{:0>2d}:{:0>2d}'.format(max_h, max_m)
        analysed_text = f'You have tracked {hours_tracked} out of {total_hours} hours.\n{max_act} has been tracked the most\n({max_tracked} hours).'
        self.data_analysed.configure(state='normal')
        self.data_analysed.delete('1.0', END)
        self.data_analysed.insert(INSERT, analysed_text)
        self.data_analysed.configure(state='disabled')
        explode = tuple([0.1 if i == max_id else 0 for i in range(len(labels))])
        figure1 = Figure(figsize=(4.5, 3), dpi=100)
        subplot1 = figure1.add_subplot(111)
        subplot1.pie(values, explode=explode, labels=labels,
                     autopct='%1.1f%%', shadow=True, startangle=90)
        subplot1.axis('equal')
        pie1 = FigureCanvasTkAgg(figure1, self.tab3)
        pie1.get_tk_widget().place(relx=0.14, rely=0.13)
        pie1.draw()
        toolbar = NavigationToolbar2Tk(pie1, self.tab3)
        toolbar.place(relx=0.14, rely=0.6)

    def save_data(self):
        """
        Save the data that is analysed.
        """
        os.chdir(image_file)
        folder_name = image_file.split('\\')[-1]
        image_name = '{}.png'.format(self.d1.replace('-', '_'))
        left = self.master.winfo_rootx()
        upper = self.master.winfo_rooty()
        right = left + self.master.winfo_width()
        lower = upper + self.master.winfo_height()
        bb = (left, upper, right, lower)
        # grab = ImageGrab.grab(bbox = bb) #Uncomment this out and run this.
        grab = ImageGrab.grab()
        grab.save(image_name)
        messagebox.showinfo(
            "Saved", "Your report has been saved as an image in the folder {}!".format(folder_name))


def main():
    """
    Create an obj and tk window and except KeyboardInterrupt
    """
    root = Tk()
    obj = TimeLogger(root)
    root.mainloop()


if __name__ == '__main__':
    main()
